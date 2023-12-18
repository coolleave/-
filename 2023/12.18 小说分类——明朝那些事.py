import re
import openpyxl


def handle_novel_mc():
    # 打开Excel文件
    workbook = openpyxl.Workbook()
    worksheet = workbook.active  # 设置为当前的活动工作表，对这个工作表进行操作
    worksheet.title = 'Novel Chapters'  # 设置工作表名称

    # 打开小说文本文件
    with open('明朝那些事儿.txt', 'r', encoding='utf-8') as file:
        lines = file.readlines()

    # 将文本按章节保存到Excel中
    chapter_index = 1
    chapter_title = ''
    chapter_content = ''
    for line in lines:
        line = line.strip()
        # print(line)
        if line.startswith('【'):
            # 保存上一章节的内容
            if chapter_title and chapter_content:
                # worksheet.cell(row=chapter_index, column=1, value=chapter_title)
                chapter_content = re.sub(r'\n', r"\\n", chapter_content)
                worksheet.cell(row=chapter_index, column=1, value=chapter_content)
                # 写入 row 行（章节数）  column 列 value 值
                chapter_index += 1
                chapter_title = ''
                chapter_content = ''

            # 提取新章节的标题
            chapter_title = line
        else:
            # 拼接章节内容
            chapter_content += line + '\n'
            print(chapter_content)

    # 保存最后一章节的内容
    if chapter_title and chapter_content:
        # worksheet.cell(row=chapter_index, column=1, value=chapter_title)
        chapter_content = re.sub(r'\n', r"\\n", chapter_content)
        # print(chapter_content)
        worksheet.cell(row=chapter_index, column=1, value=chapter_content)

    # 保存Excel文件
    workbook.save('历史.xlsx')


if __name__ == '__main__':
    handle_novel_mc()
    print("over!")
