from openpyxl import load_workbook
import win32com.client
import os


# 加载表格并读取数据
def read_excel(filename):
    wb = load_workbook(filename)
    # 设置当前表格
    ws = wb.active

    data = []
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        if len(row) >= 2:  # 确保行中至少有两列数据
            data.append((row[1], row[2]))  # 获取第二列和第三列的数据
        else:
            print(f"跳过行：{row}，因为列数不足2")
    return data[1:]  # 去除第0行空行


# 替换word信息
def replace_multiple_text_in_doc(file_path, replacements):

    word = win32com.client.Dispatch("Word.Application")
    word.Visible = False
    doc = word.Documents.Open(file_path)

    # 遍历所有要替换的文本
    for old_text, new_text in replacements.items():
        # 替换文档中的文本
        for story in doc.StoryRanges:
            find_object = story.Find
            find_object.Text = old_text
            found = find_object.Execute()
            while found:
                story.Text = story.Text.replace(old_text, new_text)
                found = find_object.Execute()
    # 提取地址+名称
    new_path = 'D:\\python\\项目\\4.6_1\\4.16\\H\\HH\\'+f'{data_handle[i][0]}-{data_handle[i][1]}-V1.doc'

    # 保存到指定地址
    doc.SaveAs(new_path)
    doc.Close()
    word.Quit()


# 列表去重，相同的人不同版本只保留第1份
def keep_first_with_same_prefix(lst):
    result = {}
    for item in lst:
        prefix = item[:20]  # 获取前28位作为前缀
        if prefix not in result:
            result[prefix] = item  # 如果前缀不存在于结果中，则将当前元素加入结果
    return list(result.values())


# 找到所有的word文件
def open_word_files(folder_path):
    name_lst = []
    word_files = [f for f in os.listdir(folder_path) if f.endswith('.doc')]


    for file_name in word_files:
        file_path = os.path.join(folder_path, file_name)

        # print(f"{file_path}")
        name_lst.append(file_path)
    return name_lst


if __name__ == "__main__":

    excel_filename = '2.xlsx'

    data = read_excel(excel_filename)
    data_handle = data[1:]

    # 提取名称
    name_lst = open_word_files('./')

    # 列表去重
    new_list = keep_first_with_same_prefix(name_lst)
    # 循环列表索引
    for i in range(71):
        # 要替换的文本字典，键是旧文本，值是新文本
        old_order = new_list[i][7:17]  # 提取旧学号
        print('old_order', old_order)
        old_name = new_list[i][13:16]  # 提取旧名字
        print('oldname', old_name)
        # 替换写成字典形式
        replacements = {
            old_order: data_handle[i][0],
            old_name: data_handle[i][1],
            # 添加更多的替换对...
        }

        # 拼接路径
        path = 'D:\\python\\项目\\4.6_1\\4.16\\H\\H2\\' + new_list[i][2:]
        # 替换文件路径
        doc_file_path = path

        # 执行替换
        replace_multiple_text_in_doc(doc_file_path, replacements)






